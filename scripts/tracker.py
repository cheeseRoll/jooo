"""Application tracker — tracker/applications.xlsx.

  python3 scripts/tracker.py init
  python3 scripts/tracker.py mark-applied <job_id>
  python3 scripts/tracker.py set-contacts <job_id> --contacts "Name <a@b.com> (CFO); ..." [--note "..."]
  python3 scripts/tracker.py set-status <job_id> <status>
  python3 scripts/tracker.py list

Row lifecycle: applied → "send cover mail" → "cover draft in Gmail" → mailed →
interview → offer / rejected / no response.
"""
import argparse
import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "tracker" / "applications.xlsx"
JOBS_DB = ROOT / "data" / "jobs.json"

HEADERS = ["Job ID", "Company", "Role", "Location", "Fit", "Job URL", "Applied on",
           "Status", "Contacts (best people to email)", "Cover letter", "Notes"]
WIDTHS = [16, 22, 32, 20, 6, 40, 12, 22, 45, 60, 30]


def _open():
    if not XLSX.exists():
        init()
    return load_workbook(XLSX)


def init():
    XLSX.parent.mkdir(exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(HEADERS)
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    wb.save(XLSX)
    print(f"created {XLSX.relative_to(ROOT)}")


def _find_row(ws, job_id):
    for row in ws.iter_rows(min_row=2):
        if row[0].value == job_id:
            return row[0].row
    return None


def mark_applied(job_id):
    db = json.loads(JOBS_DB.read_text())
    job = db["jobs"].get(job_id)
    if job is None:
        raise SystemExit(f"unknown job id {job_id}")
    job["status"] = "applied"
    job["applied_on"] = date.today().isoformat()
    JOBS_DB.write_text(json.dumps(db, indent=1, ensure_ascii=False))

    wb = _open()
    ws = wb.active
    if _find_row(ws, job_id):
        print("already in tracker")
        return
    ws.append([job_id, job["company"], job["title"], job["location"],
               job.get("fit_score"), job["url"], job["applied_on"],
               "send cover mail", "", "", ""])
    r = ws.max_row
    ws.cell(row=r, column=8).fill = PatternFill("solid", fgColor="FFF2CC")
    for c in (9, 10, 11):
        ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(XLSX)
    print(f"tracked: {job['company']} — {job['title']} (status: send cover mail)")


def set_contacts(job_id, contacts, note=None):
    wb = _open()
    ws = wb.active
    r = _find_row(ws, job_id)
    if r is None:
        raise SystemExit(f"{job_id} not in tracker — run mark-applied first")
    ws.cell(row=r, column=9, value=contacts)
    ws.cell(row=r, column=10, value=note or "cover draft in Gmail")
    ws.cell(row=r, column=8, value="cover draft in Gmail")
    ws.cell(row=r, column=8).fill = PatternFill("solid", fgColor="D9EAD3")
    wb.save(XLSX)
    print(f"contacts saved for {job_id}")


def set_status(job_id, status):
    wb = _open()
    ws = wb.active
    r = _find_row(ws, job_id)
    if r is None:
        raise SystemExit(f"{job_id} not in tracker")
    ws.cell(row=r, column=8, value=status)
    wb.save(XLSX)
    print(f"{job_id} → {status}")


def list_rows():
    wb = _open()
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        print(f"{row[0]}  {row[1][:20]:20}  {str(row[2])[:30]:30}  {row[7]}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    sub.add_parser("init")
    sub.add_parser("list")
    p = sub.add_parser("mark-applied"); p.add_argument("job_id")
    p = sub.add_parser("set-contacts"); p.add_argument("job_id")
    p.add_argument("--contacts", required=True); p.add_argument("--note")
    p = sub.add_parser("set-status"); p.add_argument("job_id"); p.add_argument("status")
    a = ap.parse_args()
    {"init": init, "list": list_rows,
     "mark-applied": lambda: mark_applied(a.job_id),
     "set-contacts": lambda: set_contacts(a.job_id, a.contacts, a.note),
     "set-status": lambda: set_status(a.job_id, a.status)}[a.cmd]()
